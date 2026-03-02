from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from io import BytesIO
from datetime import datetime

class ExportService:
    @staticmethod
    def generate_excel(simulation_result: dict, loan_data: dict) -> BytesIO:
        """Generate Excel file with amortization schedule"""
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Header
        ws_summary['A1'] = 'Loan Freedom Planner - Summary'
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary.merge_cells('A1:D1')
        
        row = 3
        ws_summary[f'A{row}'] = 'Loan Details'
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ws_summary[f'A{row}'] = 'Loan Amount:'
        ws_summary[f'B{row}'] = f"₹{loan_data['loan_amount']:,.2f}"
        row += 1
        
        ws_summary[f'A{row}'] = 'Interest Rate:'
        ws_summary[f'B{row}'] = f"{loan_data['interest_rate']}%"
        row += 1
        
        ws_summary[f'A{row}'] = 'Tenure:'
        ws_summary[f'B{row}'] = f"{loan_data['tenure_years']} years"
        row += 2
        
        # Comparison table
        ws_summary[f'A{row}'] = 'Metric'
        ws_summary[f'B{row}'] = 'Original'
        ws_summary[f'C{row}'] = 'Optimized'
        ws_summary[f'D{row}'] = 'Savings'
        
        for col in ['A', 'B', 'C', 'D']:
            ws_summary[f'{col}{row}'].font = Font(bold=True)
            ws_summary[f'{col}{row}'].fill = PatternFill(start_color='064E3B', end_color='064E3B', fill_type='solid')
            ws_summary[f'{col}{row}'].font = Font(color='FFFFFF', bold=True)
        
        row += 1
        
        original = simulation_result['original']
        optimized = simulation_result['optimized']
        savings = simulation_result['savings']
        
        ws_summary[f'A{row}'] = 'EMI'
        ws_summary[f'B{row}'] = f"₹{original['emi']:,.2f}"
        ws_summary[f'C{row}'] = f"₹{optimized['emi']:,.2f}"
        ws_summary[f'D{row}'] = '-'
        row += 1
        
        ws_summary[f'A{row}'] = 'Tenure (months)'
        ws_summary[f'B{row}'] = original['actual_tenure']
        ws_summary[f'C{row}'] = optimized['actual_tenure']
        ws_summary[f'D{row}'] = savings['months_saved']
        row += 1
        
        ws_summary[f'A{row}'] = 'Total Interest'
        ws_summary[f'B{row}'] = f"₹{original['total_interest']:,.2f}"
        ws_summary[f'C{row}'] = f"₹{optimized['total_interest']:,.2f}"
        ws_summary[f'D{row}'] = f"₹{savings['interest_saved']:,.2f}"
        row += 1
        
        ws_summary[f'A{row}'] = 'Total Repayment'
        ws_summary[f'B{row}'] = f"₹{original['total_repayment']:,.2f}"
        ws_summary[f'C{row}'] = f"₹{optimized['total_repayment']:,.2f}"
        ws_summary[f'D{row}'] = f"₹{original['total_repayment'] - optimized['total_repayment']:,.2f}"
        row += 1
        
        # Optimized Schedule
        ws_schedule = wb.create_sheet("Amortization Schedule")
        headers = ['Month', 'Date', 'Opening Balance', 'EMI', 'Interest', 'EMI Principal', 'Extra Principal', 'Prepayment Charge', 'Total Principal', 'Closing Balance']
        ws_schedule.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws_schedule.cell(1, col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='064E3B', end_color='064E3B', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        
        for entry in simulation_result['optimized_schedule']:
            ws_schedule.append([
                entry['month'],
                entry['date'],
                f"₹{entry['opening_balance']:,.2f}",
                f"₹{entry['emi']:,.2f}",
                f"₹{entry['interest']:,.2f}",
                f"₹{entry['emi_principal']:,.2f}",
                f"₹{entry['extra_principal']:,.2f}",
                f"₹{entry['prepayment_charge']:,.2f}",
                f"₹{entry['total_principal']:,.2f}",
                f"₹{entry['closing_balance']:,.2f}"
            ])
        
        # Adjust column widths
        for i, col in enumerate(ws_summary.columns, 1):
            col_letter = chr(64 + i)  # A, B, C, D...
            ws_summary.column_dimensions[col_letter].width = 20
        
        for i, col in enumerate(ws_schedule.columns, 1):
            col_letter = chr(64 + i)  # A, B, C, D...
            ws_schedule.column_dimensions[col_letter].width = 18
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def generate_pdf(simulation_result: dict, loan_data: dict) -> BytesIO:
        """Generate PDF summary report"""
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#064E3B'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        title = Paragraph("Loan Freedom Planner - Summary Report", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))
        
        # Loan details
        loan_details_data = [
            ['Loan Details', ''],
            ['Loan Amount', f"₹{loan_data['loan_amount']:,.2f}"],
            ['Interest Rate', f"{loan_data['interest_rate']}% p.a."],
            ['Tenure', f"{loan_data['tenure_years']} years"],
            ['Loan Type', loan_data.get('loan_type', 'Home')],
        ]
        
        loan_table = Table(loan_details_data, colWidths=[3*inch, 3*inch])
        loan_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#064E3B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        
        elements.append(loan_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Comparison table
        original = simulation_result['original']
        optimized = simulation_result['optimized']
        savings = simulation_result['savings']
        
        comparison_data = [
            ['Metric', 'Original', 'Optimized', 'Savings'],
            ['EMI', f"₹{original['emi']:,.2f}", f"₹{optimized['emi']:,.2f}", '-'],
            ['Tenure (months)', str(original['actual_tenure']), str(optimized['actual_tenure']), str(savings['months_saved'])],
            ['Total Interest', f"₹{original['total_interest']:,.2f}", f"₹{optimized['total_interest']:,.2f}", f"₹{savings['interest_saved']:,.2f}"],
            ['Total Repayment', f"₹{original['total_repayment']:,.2f}", f"₹{optimized['total_repayment']:,.2f}", f"₹{original['total_repayment'] - optimized['total_repayment']:,.2f}"],
        ]
        
        comparison_table = Table(comparison_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        comparison_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#064E3B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F4')]),
        ]))
        
        elements.append(comparison_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Advisory note
        if savings['interest_saved'] > 0:
            advisory_style = ParagraphStyle(
                'Advisory',
                parent=styles['Normal'],
                fontSize=11,
                textColor=colors.HexColor('#064E3B'),
                spaceAfter=12,
                leftIndent=20,
                rightIndent=20
            )
            
            advisory_text = f"""<b>Advisory:</b> By implementing the optimized repayment strategy, you can save ₹{savings['interest_saved']:,.2f} in interest 
            and become loan-free {savings['years_saved']} years earlier. This is a significant financial achievement that puts you on the path to true financial freedom."""
            
            advisory = Paragraph(advisory_text, advisory_style)
            elements.append(advisory)
        
        # Investment comparison
        if simulation_result['investment_analysis']['enabled']:
            elements.append(Spacer(1, 0.2*inch))
            
            inv_analysis = simulation_result['investment_analysis']
            
            inv_data = [
                ['Prepay vs Invest Analysis', ''],
                ['Total Prepayments', f"₹{inv_analysis['total_prepayments']:,.2f}"],
                ['Investment Return Rate', f"{inv_analysis['investment_return_rate']}% p.a."],
                ['Investment Future Value', f"₹{inv_analysis['investment_future_value']:,.2f}"],
                ['Interest Saved by Prepaying', f"₹{inv_analysis['interest_saved_by_prepaying']:,.2f}"],
                ['Difference', f"₹{inv_analysis['difference']:,.2f}"],
                ['Recommendation', inv_analysis['recommendation']],
            ]
            
            inv_table = Table(inv_data, colWidths=[3*inch, 3*inch])
            inv_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D97706')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ]))
            
            elements.append(inv_table)
        
        # Footer
        elements.append(Spacer(1, 0.5*inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        footer = Paragraph(f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')} | Loan Freedom Planner", footer_style)
        elements.append(footer)
        
        doc.build(elements)
        output.seek(0)
        return output
